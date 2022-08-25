from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import time

from selenium.webdriver.remote.webelement import WebElement

# Driver para ingresar a la página inicial de EPI
driver = webdriver.Chrome(executable_path=r"C:\Users\luis.pizarro.a\PycharmProjects\Auto_Descarte_Eno\chromedriver.exe")
driver.get("https://epivigila.minsal.cl/index.php/administracion/ver-formulario/50")
time.sleep(3)  # tiempo de carga
# Cargar el excel que tiene los datos de usuario (por ahora no se usa)
filesheet = "./User_Pass.xlsx"
wb = load_workbook(filesheet)
# hojas = wb.get_sheet_names()
# print(hojas)
# hojas = wb.sheetnames
# print(hojas)
# nombres = wb["Hoja1"]
# wb.close()

# LOGIN EN EPIVIGILA CON USER Y PASS++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
print("Ingresando Usuario.......")
usuario = driver.find_element(By.NAME, "rut_usuario")
usuario.send_keys("192572843")
print("Usuario ingresado exitosamente")
print("Ingresando Contraseña.......")
contrasena = driver.find_element(By.NAME, "password")
contrasena.send_keys("Minsal2021")
print("Contraseña ingresada exitosamente")
print("Iniciando sesión.......")
loginbutton = driver.find_element(By.XPATH, "//input[@type='submit'][@type='submit']")
loginbutton.click()
time.sleep(5)
print("Esperando captcha.......")
time.sleep(3)
print("Ingresando a la notificación....")
