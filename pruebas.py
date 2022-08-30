from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import time

from selenium.webdriver.remote.webelement import WebElement

# Driver para ingresar a la página inicial de EPI
driver = webdriver.Chrome(executable_path=r"C:\Users\luis.pizarro.a\PycharmProjects\Auto_Descarte_Eno\chromedriver.exe")
driver.get("https://epivigila.minsal.cl/index.php/administracion/ver-formulario/50")
time.sleep(3)  # tiempo de carga
# Cargar el excel que tiene los datos de usuario (por ahora no se usa)
filesheet = "./User_Pass.xlsx"
wb = load_workbook(filesheet)
# hojas = wb.get_sheet_names()
# print(hojas)
# hojas = wb.sheetnames
# print(hojas)
# nombres = wb["Hoja1"]
# wb.close()

# LOGIN EN EPIVIGILA CON USER Y PASS++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
print("Ingresando Usuario.......")
usuario = driver.find_element(By.NAME, "rut_usuario")
usuario.send_keys("192572843")
print("Usuario ingresado exitosamente")
print("Ingresando Contraseña.......")
contrasena = driver.find_element(By.NAME, "password")
contrasena.send_keys("Minsal2021")
print("Contraseña ingresada exitosamente")
print("Iniciando sesión.......")
loginbutton = driver.find_element(By.XPATH, "//input[@type='submit'][@type='submit']")
loginbutton.click()
time.sleep(5)
print("Esperando captcha.......")
time.sleep(3)
print("Ingresando a la notificación....")

"""# PASOS PARA INGRESAR AL FORMULARIO PARA LA CREACIÓN DE UNA NOTIFICACIÓN+++++++++++++++++++++++++++++++++++++++++++++
digitar = driver.find_element(By.LINK_TEXT, 'Digitar')
digitar.click()
coronavirus = driver.find_element(By.PARTIAL_LINK_TEXT, "Coronavirus")
coronavirus.click()"""

# ENTRAR A LA NOTIFICACIÓN POR UN FOLIO /50/FOLIO+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
notificacion = "https://epivigila.minsal.cl/index.php/administracion/ver-formulario/50/"
folio = 26228869
folio = folio + 50000
enlacenotificacion = notificacion + str(folio)
# print(enlacenotificacion)  # Prueba para saber si funciona la unión del link mas el folio
print("Ingresando a la sección 1...........")
# SECCIÓN 1 IDENTIFICACIÓN DEL CASO+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# -Ir a la sección Identificación del caso------------------------------------------------------------------------------
driver.get(enlacenotificacion)
time.sleep(3)

# SE DEBE INGRESAR DESCONOCIDO EN OCUPACIÓN PRINCIPAL  Y RUBRO: SI ESTOS ESTÁN VACIOS
ocupacion_principal = driver.find_element(By.XPATH, "//*[@id='actividad_laboral_declarada']")
time.sleep(1)
ocupacion_principal.send_keys("DESCONOCIDO")
time.sleep(1)
rubro = driver.find_element(By.XPATH, "//*[@id='rubro_trabajo']")
time.sleep(1)
rubro.send_keys("DESCONOCIDO")
time.sleep(1)
# SECCIÓN 2 ANTECEDENTES CLINICOS Y EPIDEMIOLOGICOS+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# -Ir a la sección Antecedentes clínicos y epidemiológicos------------------------------------------------------
print("Ingresando a la sección 2del registro: ")
seccion_2 = driver.find_element(By.PARTIAL_LINK_TEXT, 'Antecedentes clínicos y epidemiológicos')
seccion_2.click()
time.sleep(3)
# Ingreso de la fecha de primera consulta
print("Ingresando fecha de primera consulta......")
time.sleep(2)
# fecha_primera_c = driver.find_element(By.ID, 'fecha_primera_consulta')
# fecha_primera_c.send_keys(fecha_p_c)
# //*[@id="id_institucion_primera_consulta"]


"""est_salud_antc_cli =driver.find_element(By.XPATH, "//*[@id='select2-id_institucion_primera_consulta-container']")
time.sleep(1)
est_salud_antc_cli.click()
time.sleep(1)
mov_lis = driver.find_element(By.XPATH, "//*[@id='select2-id_institucion_primera_consulta-results']")"""
driver.execute_script("document.getElementsByName('id_institucion_primera_consulta')[0].value='Ejercicio libre'")
# mov_lis.find_element(By.XPATH, "//*[@id='select2-id_institucion_primera_consulta-results']/li[7]")
